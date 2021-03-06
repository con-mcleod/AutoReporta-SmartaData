#!/usr/bin/python3

import sys, csv, os, sqlite3, re
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from sklearn import linear_model
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.utils import get_column_letter
from helper_fns import *

##############################
#                            #
# PROGRAM EXECUTION          #
#                            #
##############################

if (len(sys.argv) != 2):
	print ("Usage: python3 generator.py [output_file]")
	exit(1)

output_file = sys.argv[1]

if(os.path.exists(output_file)):
	os.remove(output_file)

##############################
#                            #
# store data in lists        #
#                            #
##############################

SMIs = get_all_SMIs()
num_rows = len(SMIs)+2
dates = get_all_dates()

##############################
#                            #
# populating xlsx file       #
#                            #
##############################

wb = Workbook()

# initiate daily data worksheet
ws = wb.active
ws.title = "Daily Report"

# initiate adjusted daily data worksheet
ws2 = wb.create_sheet()
ws2.title = "Daily Unadjusted"

# initiate encompass daily data worksheet
ws3 = wb.create_sheet()
ws3.title = "Daily Adjusted"

# initiate summary data worksheet
ws4 = wb.create_sheet()
ws4.title = "Summary Stats"

# merge Summary Stats headings and center
ws4.merge_cells('A1:G1')
ws4.merge_cells('H1:K1')
for cell in ws4['1:1']:
	cell.alignment = Alignment(horizontal='center')

# bold the top 2 rows and center the top row
bolded_font = Font(bold=True)
for i in range(1,24+5*len(dates)+1):
	ws.cell(row=1, column=i).font = bolded_font
	ws.cell(row=2, column=i).font = bolded_font
	ws4.cell(row=1, column=i).font = bolded_font
	ws4.cell(row=2, column=i).font = bolded_font
for cell in ws['1:1']:
	cell.alignment = Alignment(horizontal='center')

# apply conditional format to flag under/overperforming performance
redFill = PatternFill(start_color='FA5858', end_color='FA5858', fill_type='solid')
greenFill = PatternFill(start_color='9Afe2e', end_color='9Afe2e', fill_type='solid')
noFill = PatternFill(fill_type=None)
for i in range(0,len(dates)*5, 5):
	actual_perf = get_column_letter(26+i)
	new_perf = get_column_letter(26+i+3)
	actual_perf = str(actual_perf) + "3:" + str(actual_perf) + str(num_rows-1)
	new_perf = str(new_perf) + "3:" + str(new_perf) + str(num_rows-1)
	ws.conditional_formatting.add(actual_perf,CellIsRule(operator='lessThan', formula=['.7'], fill=redFill))
	ws.conditional_formatting.add(actual_perf,CellIsRule(operator='greaterThan', formula=['1.3'], fill=greenFill))
	ws.conditional_formatting.add(new_perf,CellIsRule(operator='lessThan', formula=['.7'], fill=redFill))
	ws.conditional_formatting.add(new_perf,CellIsRule(operator='greaterThan', formula=['1.3'], fill=greenFill))

# apply conditional formatting to encompass daily performance sheet
for i in range(0,len(dates)):
	adj_perf = get_column_letter(13+i)
	adj_perf = str(adj_perf) + "3:" + str(adj_perf) + str(num_rows-1)
	ws2.conditional_formatting.add(adj_perf,CellIsRule(operator='lessThan', formula=['.7'], fill=redFill))
	ws2.conditional_formatting.add(adj_perf,CellIsRule(operator='greaterThan', formula=['1.3'], fill=greenFill))

# apply conditional formatting to adjusted daily performance sheet
for i in range(0,len(dates)):
	adj_perf = get_column_letter(13+i)
	adj_perf = str(adj_perf) + "3:" + str(adj_perf) + str(num_rows-1)
	# ws3.conditional_formatting.add(adj_perf,CellIsRule(operator='lessThan', formula=['.7'], fill=redFill))
	ws3.conditional_formatting.add(adj_perf,CellIsRule(operator='greaterThan', formula=['1.3'], fill=greenFill))
	

# apply conditional formatting to summary stats sheet
for i in range(1, 20):
	to_format = get_column_letter(i)
	to_format = str(to_format) + "3:" + str(to_format) + str(num_rows-1)
	if i == 8:
		ws4.conditional_formatting.add(to_format,CellIsRule(operator='lessThan', formula=['.7'], fill=redFill))
		ws4.conditional_formatting.add(to_format,CellIsRule(operator='greaterThan', formula=['1.3'], fill=greenFill))
	elif i == 9:
		ws4.conditional_formatting.add(to_format,ColorScaleRule(start_type='min', start_color='FA5858', mid_type='percentile', mid_value=50, mid_color='F4fa58', end_type='max', end_color='9Afe2e'))
	elif i == 10:
		ws4.conditional_formatting.add(to_format,CellIsRule(operator='greaterThan', formula=['.2'], fill=redFill))
	elif i == 11:
		ws4.conditional_formatting.add(to_format,CellIsRule(operator='greaterThan', formula=['5'], fill=redFill))
	elif i == 12:
		ws4.conditional_formatting.add(to_format,ColorScaleRule(start_type='min', start_color='FA5858', mid_type='percentile', mid_value=50, mid_color='F4fa58', end_type='max', end_color='9Afe2e'))

# merge and center daily data headings
ws.merge_cells('A1:L1')
ws.merge_cells('M1:X1')
for i in range(0,len(dates)*5, 5):
	merge_lhs = get_column_letter(25+i)
	merge_rhs = get_column_letter(25+i+4)
	to_merge = str(merge_lhs) + "1:" + str(merge_rhs) + "1"
	ws.merge_cells(to_merge)

# left border style
leftBorder = Border(left=Side(style='thin'))

# begin populating the worksheets
row_count = 0
for SMI in SMIs:

	col_count = 0

	# Handle headings across both sheets
	if (row_count == 0):
		# Salesforce and forecast headings
		ws.cell(row=row_count+1, column=col_count+1).value = "Salesforce Details"
		ws.cell(row=row_count+1, column=col_count+13).value = "Daily Forecast"

		ws.cell(row=row_count+2, column=col_count+1).value = "SMI"
		ws.cell(row=row_count+2, column=col_count+2).value = "Ref No"
		ws.cell(row=row_count+2, column=col_count+3).value = "ECS"
		ws.cell(row=row_count+2, column=col_count+4).value = "Installer"
		ws.cell(row=row_count+2, column=col_count+5).value = "PVsize"
		ws.cell(row=row_count+2, column=col_count+6).value = "Panel Brand"
		ws.cell(row=row_count+2, column=col_count+7).value = "Address"
		ws.cell(row=row_count+2, column=col_count+8).value = "State"
		ws.cell(row=row_count+2, column=col_count+9).value = "Site Status"
		ws.cell(row=row_count+2, column=col_count+10).value = "Install Date"
		ws.cell(row=row_count+2, column=col_count+11).value = "Supply Date"
		ws.cell(row=row_count+2, column=col_count+12).value = "Export Control"
		ws.cell(row=row_count+2, column=col_count+13).value = "Jan"
		ws.cell(row=row_count+2, column=col_count+14).value = "Feb"
		ws.cell(row=row_count+2, column=col_count+15).value = "Mar"
		ws.cell(row=row_count+2, column=col_count+16).value = "Apr"
		ws.cell(row=row_count+2, column=col_count+17).value = "May"
		ws.cell(row=row_count+2, column=col_count+18).value = "Jun"
		ws.cell(row=row_count+2, column=col_count+19).value = "Jul"
		ws.cell(row=row_count+2, column=col_count+20).value = "Aug"
		ws.cell(row=row_count+2, column=col_count+21).value = "Sep"
		ws.cell(row=row_count+2, column=col_count+22).value = "Oct"
		ws.cell(row=row_count+2, column=col_count+23).value = "Nov"
		ws.cell(row=row_count+2, column=col_count+24).value = "Dec"

		# Salesforce and summary stats headings
		ws4.cell(row=row_count+1, column=col_count+1).value = "Salesforce Details"
		ws4.cell(row=row_count+1, column=col_count+8).value = "Summary Stats"
		ws4.cell(row=row_count+2, column=col_count+1).value = "SMI"
		ws4.cell(row=row_count+2, column=col_count+2).value = "ECS"
		ws4.cell(row=row_count+2, column=col_count+3).value = "Address"
		ws4.cell(row=row_count+2, column=col_count+4).value = "State"
		ws4.cell(row=row_count+2, column=col_count+5).value = "Site Status"
		ws4.cell(row=row_count+2, column=col_count+6).value = "PV Size"
		ws4.cell(row=row_count+2, column=col_count+7).value = "Export Control"
		ws4.cell(row=row_count+2, column=col_count+8).value = "Enc Perf"
		ws4.cell(row=row_count+2, column=col_count+9).value = "Ave Adj Perf"
		ws4.cell(row=row_count+2, column=col_count+10).value = "Perf variance"
		ws4.cell(row=row_count+2, column=col_count+11).value = "Site off #days"
		ws4.cell(row=row_count+2, column=col_count+12).value = "Perf & Solar-rad Correlation"
		ws4.cell(row=row_count+2, column=col_count+13).value = "Closest stn"
		ws4.cell(row=row_count+2, column=col_count+14).value = "Distance from stn"

		# Daily data headings
		for date in dates:
			date = str(date)
			date = re.sub(r'(\(|\))', '', date)
			ws.cell(row=row_count+1, column=col_count+25).value = date
			ws.cell(row=row_count+2, column=col_count+25).value = "Actual Gen"
			ws.cell(row=row_count+2, column=col_count+26).value = "Curr Perf"
			ws.cell(row=row_count+2, column=col_count+27).value = "Solar rad"
			ws.cell(row=row_count+2, column=col_count+28).value = "Temp max"
			ws.cell(row=row_count+2, column=col_count+29).value = "Adjusted Perf"
			col_count += 5

		row_count += 1

	else:

		# series of functions to grab the relevant daily data
		SMI_details = get_SMI_details(SMI[0])
		SMI_monthly_forecast = get_SMI_forecast(SMI[0])
		SMI_daily_forecast = monthly_to_daily(SMI_monthly_forecast)

		SMI_daily_gen = get_SMI_daily_gen(SMI[0])
		SMI_daily_perf = get_daily_perf(SMI, SMI_daily_gen, SMI_daily_forecast, dates)
		average_perf = get_ave_perf(SMI_daily_perf)
		site_off_count = get_site_off(SMI_daily_gen)
		perf_variance = np.var(SMI_daily_perf)

		stn_checker = get_SMI_weather_stn(SMI[0])

		if (stn_checker):

			SMI_weather_stn = get_SMI_weather_stn(SMI[0])[0][0]

			stn_solar_perf = get_stn_solar_perf(SMI_weather_stn, "solar")
			stn_temp_perf = get_stn_solar_perf(SMI_weather_stn, "temp")
			relevant_solar_perf = filter_weather(stn_solar_perf, dates)
			relevant_temp_perf = filter_weather(stn_temp_perf, dates)

			solar_vals = get_weather_vals(relevant_solar_perf, len(dates))
			temp_vals = get_weather_vals(relevant_temp_perf, len(dates))

			average_solar = get_ave_condition(SMI_weather_stn, "solar")
			relevant_ave_solar = filter_ave(average_solar, dates)
			solar_cond_vs_ave = compare_cond_to_ave(solar_vals, dates, SMI_weather_stn, relevant_ave_solar)
			solar_correlation = np.corrcoef(SMI_daily_perf, solar_cond_vs_ave)
			closest_weather_stn = get_closest_stn(SMI)[0][0]
			# 1.60934 is converting mile to km
			SMI_stn_distance = int(1.60934*get_SMI_stn_dist(SMI)[0][0])

			temp_effect = get_temp_effect(temp_vals)

			temp_adjusted = temp_adjust(SMI_daily_perf, temp_effect)

			perf_x = np.array(temp_adjusted)
			solar_y = np.array(solar_cond_vs_ave)

			if not all(val==0 for val in perf_x):
				sol_slope, sol_intercept, sol_r_value, sol_p_value, sol_std_err = stats.linregress(perf_x, solar_y)
				sol_adjusted_perf = adjust_perf(temp_adjusted, solar_cond_vs_ave, sol_slope, sol_intercept, sol_p_value, sol_r_value)
				ave_adjusted = get_ave_perf(sol_adjusted_perf)

		else:
			ave_adjusted = ""
			solar_correlation = [["",""]]
			sol_adjusted_perf = create_list_of_blanks(len(dates))
			solar_cond_vs_ave = create_list_of_blanks(len(dates))
			temp_effect = create_list_of_blanks(len(dates))


		# populate and format the summary stats sheet
		for detail in SMI_details:

			ws4.cell(row=row_count+1, column=col_count+1).value = detail[0]
			ws4.cell(row=row_count+1, column=col_count+2).value = detail[2]
			ws4.cell(row=row_count+1, column=col_count+3).value = detail[6]
			ws4.cell(row=row_count+1, column=col_count+4).value = detail[7]
			ws4.cell(row=row_count+1, column=col_count+5).value = detail[8]
			ws4.cell(row=row_count+1, column=col_count+6).value = detail[4]
			ws4.cell(row=row_count+1, column=col_count+7).value = detail[11]
			ws4.cell(row=row_count+1, column=col_count+8).border = leftBorder
			ws4.cell(row=row_count+1, column=col_count+8).number_format = '0.00%'
			ws4.cell(row=row_count+1, column=col_count+8).value = average_perf
			ws4.cell(row=row_count+1, column=col_count+9).number_format = '0.00%'
			ws4.cell(row=row_count+1, column=col_count+9).value = ave_adjusted
			ws4.cell(row=row_count+1, column=col_count+10).number_format = '0.00%'
			ws4.cell(row=row_count+1, column=col_count+10).value = perf_variance
			ws4.cell(row=row_count+1, column=col_count+11).value = site_off_count
			ws4.cell(row=row_count+1, column=col_count+12).number_format = '0.00%'
			ws4.cell(row=row_count+1, column=col_count+12).value = solar_correlation[0][1]
			ws4.cell(row=row_count+1, column=col_count+13).value = closest_weather_stn
			ws4.cell(row=row_count+1, column=col_count+14).value = SMI_stn_distance

		# add Salesforce data to daily sheets
		for x in range(0, len(SMI_details)):
			for detail in SMI_details[x]:
				ws.cell(row=row_count+1, column=col_count+1).value = detail
				ws2.cell(row=row_count+1, column=col_count+1).value = detail
				ws3.cell(row=row_count+1, column=col_count+1).value = detail
				col_count += 1

		# add forecast data to daily sheet
		fcount = 0
		for df in SMI_daily_forecast:
			if fcount == 0:
				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
			ws.cell(row=row_count+1, column=col_count+1).value = df
			col_count += 1
			fcount += 1

		# populate the unadjusted performance sheet
		col3_count = col_count - 12
		for z in range(0, len(SMI_daily_gen)):
			for gen in SMI_daily_gen[z]:
				ws2.cell(row=row_count+1, column=col3_count+1).number_format = '0.00%'
				ws2.cell(row=row_count+1, column=col3_count+1).value = SMI_daily_perf[z]
				col3_count += 1

		# populate the adjusted performance sheet
		col4_count = col_count - 12
		for z in range(0, len(SMI_daily_gen)):
			for gen in SMI_daily_gen[z]:
				ws3.cell(row=row_count+1, column=col4_count+1).number_format = '0.00%'
				ws3.cell(row=row_count+1, column=col4_count+1).value = sol_adjusted_perf[z]
				if ws3.cell(row=row_count+1, column=col4_count+1).value in [None,'','None']:
					pass
				elif ws3.cell(row=row_count+1, column=col4_count+1).value < .7:
					ws3.cell(row=row_count+1, column=col4_count+1).fill = redFill

				col4_count += 1

		# add encompass and BOM data to daily sheet
		for y in range(0, len(SMI_daily_gen)):
			for gen in SMI_daily_gen[y]:

				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
				ws.cell(row=row_count+1, column=col_count+1).value = gen

				ws.cell(row=row_count+1, column=col_count+2).number_format = '0.00%'
				ws.cell(row=row_count+1, column=col_count+2).value = SMI_daily_perf[y]
				
				ws.cell(row=row_count+1, column=col_count+3).number_format = '0.00%'
				ws.cell(row=row_count+1, column=col_count+3).value = solar_cond_vs_ave[y]
				ws.cell(row=row_count+1, column=col_count+4).number_format = '0.00%'
				ws.cell(row=row_count+1, column=col_count+4).value = temp_effect[y]
				ws.cell(row=row_count+1, column=col_count+5).number_format = '0.00%'
				ws.cell(row=row_count+1, column=col_count+5).value = sol_adjusted_perf[y]

				col_count += 5

	row_count += 1

# Save the file
wb.save(output_file)
